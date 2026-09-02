"""Ingere vagas do LinkedIn coletadas manualmente (sessão ao vivo, com login
humano) e junta com o resultado do dia — mesma classificação, mesma planilha,
mesmo upload no Drive que o fluxo automático da Adzuna usa.

Por quê isso é manual: o LinkedIn não tem API pública de busca de vagas, e
scraping automatizado sem supervisão viola os termos de uso deles e arrisca
a conta. Este script não busca nada sozinho — ele só processa uma lista de
vagas que você (ou seu assistente, com você logado e presente) já extraiu
durante uma sessão de navegador ao vivo.

Uso:
    python src/ingest_linkedin.py caminho/para/linkedin_jobs.json

O JSON de entrada é uma lista de vagas no mesmo formato usado internamente
pelo restante do pipeline (veja src/search_adzuna.py `_normalize` para o
formato exato): title, company, location, description, url, source, id, etc.
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(PROJECT_ROOT / ".env")

import ssl_fix  # noqa: E402,F401
import classify  # noqa: E402
import summarize  # noqa: E402

log = logging.getLogger("ingest_linkedin")

DATA_DIR = PROJECT_ROOT / "data"
OUTPUT_DIR = PROJECT_ROOT / "output"
DESKTOP_RESULTS_DIR = PROJECT_ROOT.parent / "Vagas"


def _load_json(path: Path) -> dict | list:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_existing_xlsx(path: Path) -> list[dict]:
    """Lê de volta um vagas.xlsx já gerado hoje, se existir, pra não perder o que já tinha."""
    if not path.exists():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows)]
    jobs = []
    for values in rows:
        row = dict(zip(header, values))
        jobs.append({k: ("" if v is None else v) for k, v in row.items()})
    return jobs


def main() -> None:
    parser = argparse.ArgumentParser(description="Junta vagas do LinkedIn (coletadas ao vivo) ao resultado do dia.")
    parser.add_argument("jobs_file", type=Path, help="JSON com a lista de vagas coletadas do LinkedIn.")
    parser.add_argument("--sem-email", action="store_true", help="Não reenvia e-mail (só atualiza local + Drive).")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    run_date = date.today()
    run_date_str = run_date.isoformat()

    profile = _load_json(DATA_DIR / "profile.json")
    preferences = _load_json(DATA_DIR / "preferences.json")

    new_jobs = _load_json(args.jobs_file)
    log.info("%d vaga(s) do LinkedIn carregadas de %s.", len(new_jobs), args.jobs_file)

    desktop_dir = DESKTOP_RESULTS_DIR / run_date_str
    existing_xlsx = desktop_dir / "vagas.xlsx"
    existing_jobs = _load_existing_xlsx(existing_xlsx)
    log.info("%d vaga(s) já existentes hoje em %s.", len(existing_jobs), existing_xlsx)

    existing_urls = {j.get("url") for j in existing_jobs if j.get("url")}
    fresh_new_jobs = [j for j in new_jobs if j.get("url") not in existing_urls]
    log.info("%d vaga(s) do LinkedIn são novas (não duplicadas).", len(fresh_new_jobs))

    classified_new = classify.classify_jobs(fresh_new_jobs, profile, preferences) if fresh_new_jobs else []

    all_jobs = existing_jobs + classified_new
    n_alta = sum(1 for j in all_jobs if j.get("adequacao") == "Alta")

    markdown_summary = summarize.build_executive_summary_markdown(all_jobs, run_date)
    xlsx_content = summarize.jobs_to_xlsx(all_jobs)

    desktop_dir.mkdir(parents=True, exist_ok=True)
    (desktop_dir / "resumo.md").write_text(markdown_summary, encoding="utf-8")
    (desktop_dir / "vagas.xlsx").write_bytes(xlsx_content)

    results_dir = OUTPUT_DIR / "runs" / run_date_str
    results_dir.mkdir(parents=True, exist_ok=True)
    (results_dir / "resumo.md").write_text(markdown_summary, encoding="utf-8")
    (results_dir / "vagas.xlsx").write_bytes(xlsx_content)

    log.info("Total após juntar: %d vagas, %d de alta adequação.", len(all_jobs), n_alta)

    sheet_link = None
    try:
        import drive

        links = drive.upload_daily_results(markdown_summary, xlsx_content, run_date_str)
        sheet_link = links.get("planilha")
        log.info("Drive atualizado: %s", links)
    except Exception as exc:  # noqa: BLE001
        log.error("Falha ao subir pro Drive: %s", exc)

    if not args.sem_email:
        try:
            import mailer

            email_html = summarize.build_executive_summary_html(all_jobs, run_date, sheet_link=sheet_link)
            subject = (
                f"[Recolocação] Atualizado com LinkedIn — {n_alta} de alta adequação "
                f"— {run_date.strftime('%d/%m/%Y')}"
            )
            mailer.send_daily_summary(subject, email_html)
            log.info("E-mail de atualização enviado.")
        except Exception as exc:  # noqa: BLE001
            log.error("Falha ao enviar e-mail: %s", exc)

    print(f"Pronto: {len(fresh_new_jobs)} vaga(s) novas do LinkedIn juntadas. Total do dia: {len(all_jobs)}.")


if __name__ == "__main__":
    main()
