"""Monta o resumo executivo e a lista completa (XLSX) das vagas do dia."""
from __future__ import annotations

import io
from datetime import date

_ADEQUACAO_ORDEM = {"Alta": 0, "Média": 1, "Baixa": 2}

XLSX_COLUMNS = [
    "feedback", "adequacao", "nivel", "modalidade", "cidade", "title", "company",
    "adequacao_justificativa", "url", "source", "salary_min", "salary_max",
]

_COLUMN_WIDTHS = {
    "feedback": 10, "adequacao": 10, "nivel": 24, "modalidade": 12, "cidade": 18,
    "title": 42, "company": 26, "adequacao_justificativa": 55, "url": 42,
    "source": 10, "salary_min": 12, "salary_max": 12,
}

FEEDBACK_HELP = (
    "Preencha a coluna 'feedback' com 'bom' ou 'ruim' nas vagas que você já revisou "
    "(deixe em branco as que não olhou). Isso vai alimentar um classificador mais "
    "personalizado no futuro."
)


def _sorted_jobs(jobs: list[dict]) -> list[dict]:
    return sorted(jobs, key=lambda j: _ADEQUACAO_ORDEM.get(j.get("adequacao"), 9))


def jobs_to_xlsx(jobs: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Vagas"

    ws.append(XLSX_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for job in _sorted_jobs(jobs):
        ws.append([job.get(col, "") for col in XLSX_COLUMNS])

    for i, col in enumerate(XLSX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _COLUMN_WIDTHS.get(col, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_executive_summary_markdown(
    jobs: list[dict], run_date: date | None = None, top_n: int = 10, sheet_link: str | None = None
) -> str:
    run_date = run_date or date.today()
    ordered = _sorted_jobs(jobs)
    top_matches = [j for j in ordered if j.get("adequacao") == "Alta"][:top_n]
    if not top_matches:
        top_matches = ordered[:top_n]

    n_alta = sum(1 for j in jobs if j.get("adequacao") == "Alta")
    n_media = sum(1 for j in jobs if j.get("adequacao") == "Média")
    n_baixa = sum(1 for j in jobs if j.get("adequacao") == "Baixa")

    lines = [
        f"# Resumo de vagas — {run_date.strftime('%d/%m/%Y')}",
        "",
        f"**{len(jobs)} vagas encontradas** — {n_alta} de alta adequação, "
        f"{n_media} média, {n_baixa} baixa.",
        "",
        "## Melhores oportunidades",
        "",
    ]

    if not jobs:
        lines.append("Nenhuma vaga encontrada hoje.")
    else:
        for job in top_matches:
            lines.append(
                f"### [{job.get('title', 'Sem título')}]({job.get('url', '')}) — {job.get('company', 'Empresa não informada')}"
            )
            lines.append(
                f"- **Adequação**: {job.get('adequacao', 'n/d')} · "
                f"**Nível**: {job.get('nivel', 'n/d')} · "
                f"**Modalidade**: {job.get('modalidade', 'n/d')} · "
                f"**Cidade**: {job.get('cidade', 'n/d')}"
            )
            lines.append(f"- {job.get('adequacao_justificativa', '')}")
            lines.append("")

    lines.append("---")
    if sheet_link:
        lines.append(
            f"[Lista completa (Google Sheets)]({sheet_link}) com todas as {len(jobs)} vagas "
            "encontradas — preencha a coluna 'feedback' direto lá."
        )
    else:
        lines.append(f"Lista completa na planilha vagas.xlsx (nesta mesma pasta/anexo) com todas as {len(jobs)} vagas encontradas.")
    return "\n".join(lines)


def build_executive_summary_html(
    jobs: list[dict], run_date: date | None = None, top_n: int = 10, sheet_link: str | None = None
) -> str:
    run_date = run_date or date.today()
    ordered = _sorted_jobs(jobs)
    top_matches = [j for j in ordered if j.get("adequacao") == "Alta"][:top_n]
    if not top_matches:
        top_matches = ordered[:top_n]

    n_alta = sum(1 for j in jobs if j.get("adequacao") == "Alta")
    n_media = sum(1 for j in jobs if j.get("adequacao") == "Média")
    n_baixa = sum(1 for j in jobs if j.get("adequacao") == "Baixa")

    cards = []
    for job in top_matches:
        cards.append(f"""
        <div style="margin-bottom:16px;padding:12px;border:1px solid #ddd;border-radius:8px;">
          <h3 style="margin:0 0 4px;"><a href="{job.get('url', '')}">{job.get('title', 'Sem título')}</a> — {job.get('company', '')}</h3>
          <p style="margin:0 0 4px;color:#555;">
            <b>Adequação:</b> {job.get('adequacao', 'n/d')} &middot;
            <b>Nível:</b> {job.get('nivel', 'n/d')} &middot;
            <b>Modalidade:</b> {job.get('modalidade', 'n/d')} &middot;
            <b>Cidade:</b> {job.get('cidade', 'n/d')}
          </p>
          <p style="margin:0;">{job.get('adequacao_justificativa', '')}</p>
        </div>""")

    body = "".join(cards) if jobs else "<p>Nenhuma vaga encontrada hoje.</p>"

    if sheet_link:
        footer = f'<p><a href="{sheet_link}">Ver lista completa (Google Sheets)</a> com todas as {len(jobs)} vagas encontradas — preencha a coluna "feedback" direto lá.</p>'
    else:
        footer = f"<p>Lista completa em anexo (XLSX) com todas as {len(jobs)} vagas encontradas.</p>"

    return f"""
    <div style="font-family:Arial,sans-serif;">
      <h2>Resumo de vagas — {run_date.strftime('%d/%m/%Y')}</h2>
      <p><b>{len(jobs)} vagas encontradas</b> — {n_alta} de alta adequação, {n_media} média, {n_baixa} baixa.</p>
      <h3>Melhores oportunidades</h3>
      {body}
      <hr>
      {footer}
    </div>
    """
