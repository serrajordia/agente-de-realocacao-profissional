"""Coleta o feedback preenchido manualmente nas planilhas de vagas e acumula
um histórico local rotulado — base para, no futuro, treinar um classificador
no seu próprio gosto em vez de depender só da heurística/Claude (fase 2 da
triagem de vagas; ainda não treina nada, só acumula dado).

Uso:
    python src/collect_feedback.py                      # varre Desktop\\Vagas\\*\\vagas.xlsx
    python src/collect_feedback.py --xlsx caminho.xlsx   # só um arquivo específico

Preencha a coluna "feedback" da planilha com "bom" ou "ruim" nas vagas que
você já revisou (deixe em branco as que não olhou) antes de rodar este script.
"""
from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path

log = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DESKTOP_RESULTS_DIR = PROJECT_ROOT.parent / "Vagas"
HISTORY_FILE = PROJECT_ROOT / "data" / "feedback_history.jsonl"

VALID_FEEDBACK = {"bom", "ruim"}


def _load_history() -> dict[str, dict]:
    if not HISTORY_FILE.exists():
        return {}
    history = {}
    for line in HISTORY_FILE.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        record = json.loads(line)
        history[record["chave"]] = record
    return history


def _job_key(row: dict) -> str:
    return row.get("url") or f"{row.get('title', '')}|{row.get('company', '')}|{row.get('cidade', '')}"


def collect_from_xlsx(xlsx_path: Path, history: dict[str, dict]) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows)]

    added = 0
    for values in rows:
        row = dict(zip(header, values))
        feedback = str(row.get("feedback") or "").strip().lower()
        if feedback not in VALID_FEEDBACK:
            continue
        row = {k: ("" if v is None else v) for k, v in row.items()}
        key = _job_key(row)
        history[key] = {**row, "feedback": feedback, "chave": key, "fonte_xlsx": str(xlsx_path)}
        added += 1
    return added


def main() -> None:
    parser = argparse.ArgumentParser(description="Coleta feedback manual preenchido nas planilhas de vagas.")
    parser.add_argument("--xlsx", type=Path, help="Uma planilha específica (default: varre Desktop/Vagas/*/vagas.xlsx)")
    args = parser.parse_args()

    xlsx_files = [args.xlsx] if args.xlsx else sorted(DESKTOP_RESULTS_DIR.glob("*/vagas.xlsx"))
    if not xlsx_files:
        print(f"Nenhuma planilha de vagas encontrada em {DESKTOP_RESULTS_DIR}.")
        return

    history = _load_history()
    for xlsx_path in xlsx_files:
        collect_from_xlsx(xlsx_path, history)

    HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    with HISTORY_FILE.open("w", encoding="utf-8") as f:
        for record in history.values():
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    n_bom = sum(1 for r in history.values() if r["feedback"] == "bom")
    n_ruim = sum(1 for r in history.values() if r["feedback"] == "ruim")
    print(f"{len(history)} vaga(s) com feedback no histórico ({n_bom} bom, {n_ruim} ruim). Arquivo: {HISTORY_FILE}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    main()
